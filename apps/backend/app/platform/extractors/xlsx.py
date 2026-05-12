"""XLSX extraction via openpyxl.

Each sheet is rendered as a markdown-ish table so an LLM can reason about rows
+ columns without losing structure. Very wide / very tall sheets are still
emitted in full — callers truncate downstream if needed.
"""

from __future__ import annotations

from pathlib import Path

from .base import ExtractionError, ExtractResult


class XlsxExtractor:
    extensions = ("xlsx",)

    def extract(self, path: Path) -> ExtractResult:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:  # pragma: no cover
            raise ExtractionError("openpyxl is not installed") from e

        try:
            # read_only + data_only skips formulas & formatting for speed
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise ExtractionError(f"Failed to open XLSX {path.name}: {e}") from e

        parts: list[str] = []
        sheet_names: list[str] = []
        for sheet in wb.worksheets:
            sheet_names.append(sheet.title)
            parts.append(f"### Sheet: {sheet.title}")

            rows_text: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                # Skip fully-empty rows — keeps output compact
                if any(c.strip() for c in cells):
                    rows_text.append(" | ".join(cells))
            parts.append("\n".join(rows_text) if rows_text else "(empty)")

        wb.close()
        return ExtractResult(
            text="\n\n".join(parts),
            metadata={"sheets": sheet_names},
        )
